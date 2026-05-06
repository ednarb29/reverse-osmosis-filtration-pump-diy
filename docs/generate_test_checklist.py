"""
Generate the hardware-test checklist as an .xlsx file.

Run from the repo root:
    python3 docs/generate_test_checklist.py

Writes docs/hardware_test_checklist.xlsx with:
- One row per test scenario
- A Status column with a Pending/Pass/Fail dropdown
- Notes column for free-text observations
- Frozen header row, sensible column widths, color-coded section breaks

Re-run any time the test plan changes (this script is the source of truth;
the .xlsx itself is meant to be edited per-test-run, not committed).
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# Test scenarios grouped by section. Each row is:
# (id, action, expected, what_to_listen_for)
SECTIONS = [
    ("1. First-pass smoke test", [
        ("1.1", "Power on the system",
         "Greeting beep (short + long). Serial: 'Watchdog enabled (8000 ms)'. "
         "Pump silent, all 4 valves CLICK closed.",
         "Beep pattern, relay clicks, no pump sound"),
        ("1.2", "Wait 30 seconds, do nothing",
         "Nothing happens — no spurious activity.",
         "Silence"),
        ("1.3", "Single short press",
         "Short beep. Filter cycle: flush mode → pump on → filter mode → pump runs ~15s "
         "→ 3 long beeps → pump stops → valves close (~3s).",
         "Beep, relay clicks, pump on for ~15s, 3 finish beeps"),
        ("1.4", "After 1.3, wait 15 s",
         "Deferred post-flush fires: relays click into flush mode → pump runs ~5s → "
         "pump stops → valves close.",
         "Pump-on for ~5s starting ~15s after the previous cycle ended"),
        ("1.5", "After 1.4, wait 2 minutes (auto_flush_sec)",
         "Auto-flush fires: full pre-flush + disposal + post-flush cycle "
         "(~25s of activity).",
         "Pump runs through all three phases automatically"),
    ]),
    ("2. New gestures (manual mode)", [
        ("2.1", "From idle, double-tap (two short presses within ~600 ms)",
         "Two short beeps quickly. Serial: 'Enter manual mode'. "
         "NO relay clicks, NO pump.",
         "Beeps only — system stays quiet"),
        ("2.2", "While in manual mode, single short press",
         "Short beep. Relays click to FILTER configuration (V1 + V3 open). "
         "Pump starts. NO pre-flush — pump runs immediately.",
         "One beep, ONE round of relay clicks, then continuous pump"),
        ("2.3", "Single short press again (toggle off)",
         "Short beep. Pump stops. Clean shutdown (~3s of valve sequencing). "
         "Still in manual mode.",
         "Pump stops, then bleed sequence"),
        ("2.4", "Long press (≥0.8s) to exit",
         "Long beep. Serial: 'Exit manual mode'. Back to normal idle.",
         "One long beep, no further relay activity"),
        ("2.5", "After exit, wait 2 minutes",
         "Auto-flush fires normally (manual mode no longer pauses it).",
         "Auto-flush activity"),
    ]),
    ("3. Cancellation paths", [
        ("3.1", "Start a filter (short press), wait ~5s, short press again",
         "Filter cancels, clean shutdown begins immediately. "
         "NO post-flush right away — deferred starts 15s later.",
         "Cancel beep, immediate cleanup, no post-flush relay activity until 15s later"),
        ("3.2", "Start a filter, wait until pump on, then ~35–40 s in, long press (~1s)",
         "Long beep. Filter stops AND config.json gets the elapsed time written. "
         "Note: the saved value is clamped to >= 30 s (FILTER_SEC_MIN), so a press "
         "shorter than that gets bumped up to 30 s. Verify: do another short press "
         "and confirm the new cycle uses the saved duration.",
         "Long beep, immediate cleanup; saved duration printed to serial"),
        ("3.3", "During an auto-flush (from step 1.5), short-press during pump-on phase",
         "Cancels the flush cleanly. No follow-on filter cycle. Returns to idle.",
         "Cancel beep, cleanup sequence, no further pump activity"),
        ("3.4", "Start a long flush (super-long press), wait ~5 s, short press",
         "Long flush cancels cleanly. No follow-on filter cycle. Returns to idle.",
         "Cancel beep, cleanup sequence, no further pump activity"),
        ("3.5", "Start a filter (short press) from cold idle, during the pre-flush "
                "phase (first ~10 s, valves in flush mode) press short again",
         "Pre-flush is skipped — system jumps straight to the filter phase. "
         "Pump stays running through the valve transition (clean transition, "
         "ramp from flush valves -> filter valves with bleed window).",
         "Short cancel beep, valves switch to filter, then filter runs to completion"),
    ]),
    ("4. Edge cases and safety", [
        ("4.1", "Power-cycle mid-filter (yank power while pump is on)",
         "On reboot: pump OFF, valves closed (init() handles this). "
         "System resumes normal operation.",
         "Greeting beep on reboot, no pump on startup"),
        ("4.2", "Edit config.json to invalid JSON like '{xxx', reboot",
         "System still boots, uses defaults. No crash.",
         "Normal greeting beep, normal serial output"),
        ("4.3", "Hold button super-long (≥5s) from idle",
         "1-second beep, then 20-second long flush starts.",
         "Super-long beep, then long flush activity"),
        ("4.4", "While in manual mode with pump running, long press",
         "Pump stops AND manual mode exits — both at once.",
         "Long beep, cleanup, returns to normal idle"),
        ("4.5", "Disinfection rehearsal: idle → double-tap → short press for ~10s "
                "→ short press → long press",
         "Watch for ~150–200 ml at the output during the pump-on phase. "
         "This is the actual disinfection step 2/4 motion.",
         "Measured volume in the cup matches expected"),
    ]),
    ("5. Failure modes to watch for (any test)", [
        ("5.1", "Pump runs while ALL valves are closed (deadhead)",
         "Should NEVER happen. Pump straining sound + watchdog may bail.",
         "Listen for pump struggle"),
        ("5.2", "Relays clicking rapidly back and forth",
         "Sign of a cancel-and-replace race. Should be smooth transitions only.",
         "Listen for chatter"),
        ("5.3", "Pump turns off and immediately back on",
         "Sign of the deferred post-flush race. Should have at least the "
         "bleed window between OFF and ON.",
         "Listen for tight pump cycling"),
        ("5.4", "Pump keeps running after a cycle completes",
         "Broken cleanup. Should never happen.",
         "Watch the final state of every test"),
        ("5.5", "Beep pattern wrong",
         "Short=ack, long=long-press recognized, super-long=super-long recognized, "
         "3-long=cycle finished, 2-short-quickly=double-tap recognized.",
         "Mental check on every beep"),
    ]),
]


def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Hardware Tests"

    # Header
    headers = ["#", "Action", "Expected behavior", "What to listen / watch for", "Status", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="333333")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Column widths
    widths = {"A": 6, "B": 50, "C": 55, "D": 40, "E": 12, "F": 50}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Status dropdown (applied to whole column E from row 2 down)
    dv = DataValidation(
        type="list",
        formula1='"Pending,Pass,Fail,Skip,N/A"',
        allow_blank=True,
    )
    dv.add(f"E2:E{1000}")
    ws.add_data_validation(dv)

    # Conditional fill colors per status (manual approach: just default to Pending)
    section_fill = PatternFill("solid", fgColor="D9E8F5")
    pending_fill = PatternFill("solid", fgColor="FFF6E0")

    row = 2
    for section_name, rows in SECTIONS:
        # Section header row spanning columns A:F
        ws.cell(row=row, column=1, value=section_name)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        sec_cell = ws.cell(row=row, column=1)
        sec_cell.font = Font(bold=True, size=12)
        sec_cell.fill = section_fill
        sec_cell.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[row].height = 24
        row += 1

        for tid, action, expected, listen in rows:
            ws.cell(row=row, column=1, value=tid).alignment = Alignment(vertical="top")
            ws.cell(row=row, column=2, value=action).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row, column=3, value=expected).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row, column=4, value=listen).alignment = Alignment(vertical="top", wrap_text=True)
            status_cell = ws.cell(row=row, column=5, value="Pending")
            status_cell.alignment = Alignment(vertical="top", horizontal="center")
            status_cell.fill = pending_fill
            ws.cell(row=row, column=6, value="").alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 60
            row += 1

    # Second sheet: pre-test setup
    setup = wb.create_sheet("Pre-test setup")
    setup.column_dimensions["A"].width = 6
    setup.column_dimensions["B"].width = 100

    setup_steps = [
        ("Pre-test setup — do these once before running the test matrix", True),
        ("", False),
        ("1.", "Upload main.py (and the sim/ + tests/ directories optionally) to the Pico via Thonny or mpremote."),
        ("2.", "Drop a test config.json onto the Pico's filesystem with these short timings so the suite finishes in a reasonable time:"),
        ("",   "    {"),
        ("",   '      "pre_flush_sec": 5,'),
        ("",   '      "post_flush_sec": 5,'),
        ("",   '      "long_flush_sec": 20,'),
        ("",   '      "disposal_sec": 5,'),
        ("",   '      "filter_sec": 15,'),
        ("",   '      "auto_flush_sec": 120,'),
        ("",   '      "water_clean_sec": 30,'),
        ("",   '      "deferred_post_flush_sec": 15,'),
        ("",   '      "buzzer_frequency": 1500,'),
        ("",   '      "pump_switch_delay": 1000'),
        ("",   "    }"),
        ("3.", "Connect to the Pico's USB serial console (Thonny REPL or `mpremote`). The test scenarios reference serial output for verification."),
        ("4.", "Have a measuring cup ready for scenario 4.5 (the disinfection rehearsal)."),
        ("5.", "After all tests pass, replace config.json with your production values "
               "(longer timings; keep deferred_post_flush_sec=15 if you'll use the system "
               "in a camper, or set to 180 for stationary use)."),
        ("", False),
        ("Status legend", True),
        ("",   "Pending — not yet run"),
        ("",   "Pass — observed expected behavior"),
        ("",   "Fail — did NOT match expected behavior (capture serial output in Notes)"),
        ("",   "Skip — deliberately not running this scenario"),
        ("",   "N/A — scenario doesn't apply to this hardware build"),
    ]

    for i, item in enumerate(setup_steps, start=1):
        if isinstance(item[1], bool):
            # Header / blank rows
            text, is_header = item
            if text:
                setup.cell(row=i, column=2, value=text).font = Font(bold=True, size=12)
                setup.cell(row=i, column=2).fill = section_fill
            continue
        num, text = item
        setup.cell(row=i, column=1, value=num).alignment = Alignment(vertical="top")
        setup.cell(row=i, column=2, value=text).alignment = Alignment(vertical="top", wrap_text=True)
        if "{" in text or '":' in text:
            setup.cell(row=i, column=2).font = Font(name="Menlo", size=10)
        if len(text) > 80:
            setup.row_dimensions[i].height = 30

    # Move setup tab to the front
    wb._sheets = [setup, ws]

    return wb


if __name__ == "__main__":
    import os
    out = os.path.join(os.path.dirname(__file__), "hardware_test_checklist.xlsx")
    wb = build_workbook()
    wb.save(out)
    print(f"Wrote {out}")
